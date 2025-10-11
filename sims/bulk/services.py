"""Service layer for bulk operations."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, Iterator, List, Sequence

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from sims.bulk.models import BulkOperation
from sims.logbook.models import LogbookEntry
from sims.users.models import User


@dataclass
class BulkResult:
    successes: List[dict]
    failures: List[dict]


class BulkProcessingError(Exception):
    """Raised when a bulk operation fails in transactional mode."""


class BulkService:
    def __init__(self, actor: User, chunk_size: int = 50):
        self.actor = actor
        self.chunk_size = chunk_size
        self._validate_permissions()

    def _validate_permissions(self) -> None:
        if not (
            self.actor.is_superuser or getattr(self.actor, "role", None) in {"admin", "supervisor"}
        ):
            raise PermissionDenied("Bulk operations are restricted to supervisors and admins.")

    # ------------------------------------------------------------------
    # Review and assignment operations

    def review_entries(self, entry_ids: Sequence[int], status: str) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_REVIEW)
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.status = status
                        entry.supervisor_action_at = timezone.now()
                        entry.save(update_fields=["status", "supervisor_action_at"])
                        successes.append({"id": entry.pk, "status": status})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    def assign_supervisor(self, entry_ids: Sequence[int], supervisor: User) -> BulkOperation:
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_ASSIGNMENT
        )
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.supervisor = supervisor
                        entry.save(update_fields=["supervisor"])
                        successes.append({"id": entry.pk, "supervisor": supervisor.pk})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    # ------------------------------------------------------------------
    # Bulk import

    def import_logbook_entries(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
    ) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_IMPORT)
        rows = list(_parse_rows(uploaded_file))
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False

        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            try:
                pg = User.objects.get(username=row["pg_username"], role="pg")
            except User.DoesNotExist:
                failures.append({"row": row, "error": "invalid-pg"})
                errors_triggered = True
                return
            try:
                entry_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
            except (KeyError, ValueError):
                failures.append({"row": row, "error": "invalid-date"})
                errors_triggered = True
                return
            status = row.get("status", "draft")
            payload = {
                "pg": pg,
                "case_title": row.get("case_title") or "Untitled",
                "date": entry_date,
                "status": status,
                "location_of_activity": row.get("location") or "Not specified",
                "patient_history_summary": row.get("patient_history") or "Pending summary",
                "management_action": row.get("management_action") or "Pending action",
                "topic_subtopic": row.get("topic_subtopic") or "General",
            }
            try:
                if dry_run:
                    entry = LogbookEntry(**payload)
                    entry.full_clean()
                else:
                    entry = LogbookEntry.objects.create(**payload, created_by=self.actor)
                successes.append(
                    {
                        "pg": pg.username,
                        "case_title": payload["case_title"],
                        "status": status,
                    }
                )
            except ValidationError as exc:
                failures.append({"row": row, "error": exc.message_dict})
                errors_triggered = True

        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation

        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation


def _chunked(items: Sequence[int], chunk_size: int) -> Iterator[List[int]]:
    chunk: List[int] = []
    for item in items:
        chunk.append(item)
        if len(chunk) == chunk_size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


REQUIRED_COLUMNS = {"pg_username", "case_title", "date", "status"}


def _parse_rows(uploaded_file) -> Iterator[dict]:
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    if name.endswith(".csv"):
        text_stream = (
            io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
        )
        reader = csv.DictReader(text_stream)
        _validate_headers(reader.fieldnames)
        for row in reader:
            yield {key: (value or "").strip() for key, value in row.items() if key}
    elif name.endswith(".xlsx"):
        workbook = load_workbook(stream)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        _validate_headers(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            payload = {
                headers[idx]: (value or "") for idx, value in enumerate(row) if idx < len(headers)
            }
            yield {key: str(value).strip() for key, value in payload.items()}
    else:
        raise ValidationError("Unsupported file format")


def _validate_headers(headers: Iterable[str]) -> None:
    if not headers:
        raise ValidationError("No headers found in file")
    missing = REQUIRED_COLUMNS - set(header.strip() for header in headers if header)
    if missing:
        raise ValidationError(f"Missing columns: {', '.join(sorted(missing))}")
