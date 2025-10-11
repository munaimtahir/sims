"""Services for generating PDF and Excel reports."""

from __future__ import annotations

import base64
import io
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List

from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage
from django.utils import timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from sims.analytics.services import get_accessible_users
from sims.logbook.models import LogbookEntry
from sims.reports.models import ReportTemplate, ScheduledReport
from sims.users.models import User

REPORT_ROOT = Path(settings.MEDIA_ROOT) / "reports"
REPORT_ROOT.mkdir(parents=True, exist_ok=True)
REPORT_STORAGE = FileSystemStorage(location=str(REPORT_ROOT))


@dataclass
class RenderedReport:
    filename: str
    content: bytes
    content_type: str

    def as_base64(self) -> str:
        return base64.b64encode(self.content).decode("ascii")


class ReportRenderer:
    def render_pdf(self, context: dict, filename: str) -> RenderedReport:
        output = io.BytesIO()
        document = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        subtitle_style = ParagraphStyle(
            name="Subtitle", parent=styles["Normal"], alignment=1
        )

        elements = [
            Paragraph(context.get("title", "SIMS Report"), title_style),
            Paragraph(
                f"Generated at {context.get('generated_at', timezone.now()).strftime('%d %b %Y %H:%M')}",
                subtitle_style,
            ),
            Spacer(1, 12),
        ]

        columns = context.get("columns", [])
        rows = context.get("rows", [])
        if columns and rows:
            data = [columns]
            for row in rows:
                data.append([row.get(column, "") for column in columns])
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ]
                )
            )
            elements.append(table)

        document.build(elements)
        return RenderedReport(
            filename=f"{filename}.pdf",
            content=output.getvalue(),
            content_type="application/pdf",
        )

    def render_excel(
        self, rows: List[dict], columns: Iterable[str], filename: str
    ) -> RenderedReport:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        header_font = Font(bold=True)
        for col_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_index, value=column)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(col_index)].width = max(
                15, len(column) + 2
            )
        for row_index, row in enumerate(rows, start=2):
            for col_index, column in enumerate(columns, start=1):
                cell = sheet.cell(
                    row=row_index, column=col_index, value=row.get(column)
                )
                cell.alignment = Alignment(vertical="top")
        sheet.freeze_panes = "A2"
        output = io.BytesIO()
        workbook.save(output)
        return RenderedReport(
            filename=f"{filename}.xlsx",
            content=output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class ReportService:
    def __init__(self, actor: User):
        self.actor = actor
        self.renderer = ReportRenderer()

    def generate(
        self, template: ReportTemplate, params: dict, fmt: str
    ) -> RenderedReport:
        context = self._build_context(template, params)
        filename = f"{template.slug}-{timezone.now():%Y%m%d%H%M%S}"
        if fmt == "pdf":
            return self.renderer.render_pdf(context, filename)
        if fmt == "xlsx":
            return self.renderer.render_excel(
                context["rows"], context["columns"], filename
            )
        raise ValueError("Unsupported format")

    def _build_context(self, template: ReportTemplate, params: dict) -> dict:
        if template.slug == "logbook-summary":
            return self._logbook_summary_context(params)
        context = template.default_params.copy()
        context.update(params)
        return context

    def _logbook_summary_context(self, params: dict) -> dict:
        start = params.get("start_date")
        end = params.get("end_date")
        pg_id = params.get("pg_id")
        queryset = LogbookEntry.objects.select_related("pg", "supervisor").order_by(
            "-date"
        )
        accessible = get_accessible_users(self.actor)
        queryset = queryset.filter(pg__in=accessible)
        if start:
            queryset = queryset.filter(
                date__gte=datetime.strptime(start, "%Y-%m-%d").date()
            )
        if end:
            queryset = queryset.filter(
                date__lte=datetime.strptime(end, "%Y-%m-%d").date()
            )
        if pg_id:
            queryset = queryset.filter(pg_id=pg_id)
        rows: List[dict] = []
        for entry in queryset:
            rows.append(
                {
                    "Date": entry.date.isoformat(),
                    "Postgraduate": entry.pg.get_full_name() if entry.pg else "",
                    "Supervisor": (
                        entry.supervisor.get_full_name() if entry.supervisor else ""
                    ),
                    "Case Title": entry.case_title,
                    "Status": entry.get_status_display(),
                }
            )
        context = {
            "title": "Logbook Summary",
            "generated_at": timezone.now(),
            "entries": queryset,
            "rows": rows,
            "columns": ["Date", "Postgraduate", "Supervisor", "Case Title", "Status"],
        }
        return context


class ScheduledReportRunner:
    def __init__(self, actor: User | None = None):
        self.actor = actor or User.objects.filter(role="admin").first()

    def run_due_reports(self) -> int:
        now = timezone.now()
        schedules = ScheduledReport.objects.filter(is_active=True, next_run_at__lte=now)
        count = 0
        for schedule in schedules:
            service = ReportService(self.actor or schedule.created_by)
            fmt = schedule.params.get("format", "pdf")
            try:
                report = service.generate(schedule.template, schedule.params, fmt)
                path = REPORT_STORAGE.save(report.filename, io.BytesIO(report.content))
                self._email_report(schedule, report)
                schedule.record_run(True, {"path": path})
                count += 1
            except Exception as exc:
                schedule.record_run(False, {"error": str(exc)})
        return count

    def _email_report(self, schedule: ScheduledReport, report: RenderedReport) -> None:
        recipients = [
            email.strip() for email in schedule.email_to.split(",") if email.strip()
        ]
        if not recipients:
            return
        message = EmailMessage(
            subject=f"Scheduled report: {schedule.template.name}",
            body="Please find the attached report.",
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "noreply@example.com"),
            to=recipients,
        )
        message.attach(report.filename, report.content, report.content_type)
        message.send(fail_silently=True)


__all__ = ["ReportService", "ReportRenderer", "RenderedReport", "ScheduledReportRunner"]
