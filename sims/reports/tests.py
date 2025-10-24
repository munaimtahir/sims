from __future__ import annotations

import base64
import io
from datetime import date

from django.core.management import call_command
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from sims.logbook.models import LogbookEntry
from sims.reports.models import ReportTemplate, ScheduledReport
from sims.reports.services import ReportService
from sims.users.models import User


class ReportingTests(APITestCase):
    def setUp(self) -> None:
        self.admin = User.objects.create_user(username="admin", password="testpass", role="admin")
        self.supervisor = User.objects.create_user(
            username="sup",
            password="testpass",
            role="supervisor",
            specialty="surgery",
        )
        self.pg = User.objects.create_user(
            username="pg",
            password="testpass",
            role="pg",
            specialty="surgery",
            year="1",
            supervisor=self.supervisor,
        )
        self.client.force_authenticate(self.admin)
        LogbookEntry.objects.create(
            pg=self.pg,
            case_title="Report Case",
            date=date(2024, 1, 1),
            location_of_activity="Ward",
            patient_history_summary="History",
            management_action="Action",
            topic_subtopic="Topic",
            status="approved",
            supervisor=self.supervisor,
        )
        self.template, _ = ReportTemplate.objects.get_or_create(
            slug="logbook-summary",
            defaults={
                "name": "Logbook Summary",
                "template_name": "reports/logbook_summary.html",
            },
        )

    def test_service_generates_pdf_and_excel(self) -> None:
        service = ReportService(self.admin)
        pdf_report = service.generate(self.template, {"format": "pdf"}, "pdf")
        self.assertTrue(pdf_report.content.startswith(b"%PDF"))

        excel_report = service.generate(self.template, {"format": "xlsx"}, "xlsx")
        workbook = load_workbook(io.BytesIO(excel_report.content))
        sheet = workbook.active
        self.assertEqual(sheet.cell(row=2, column=4).value, "Report Case")

    def test_generate_api_returns_base64(self) -> None:
        url = reverse("reports_api:generate")
        response = self.client.post(
            url,
            {"template_slug": self.template.slug, "format": "pdf", "params": {}},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        decoded = base64.b64decode(response.data["content"])
        self.assertTrue(decoded.startswith(b"%PDF"))

    def test_scheduled_report_runner(self) -> None:
        schedule = ScheduledReport.objects.create(
            template=self.template,
            created_by=self.admin,
            email_to="admin@example.com",
            params={"format": "pdf"},
            cron="0 6 * * *",
            next_run_at=timezone.now(),
        )
        call_command("run_scheduled_reports")
        schedule.refresh_from_db()
        self.assertIsNotNone(schedule.last_run_at)
        self.assertTrue(schedule.last_result.get("success"))

    def test_pdf_generation_with_empty_data(self) -> None:
        """Test PDF generation with no entries."""
        service = ReportService(self.admin)
        LogbookEntry.objects.all().delete()
        pdf_report = service.generate(self.template, {}, "pdf")
        self.assertTrue(pdf_report.content.startswith(b"%PDF"))
        self.assertIn("pdf", pdf_report.filename)

    def test_xlsx_generation_with_filters(self) -> None:
        """Test XLSX generation with date filters."""
        service = ReportService(self.admin)
        params = {
            "start_date": "2024-01-01",
            "end_date": "2024-12-31",
        }
        excel_report = service.generate(self.template, params, "xlsx")
        workbook = load_workbook(io.BytesIO(excel_report.content))
        sheet = workbook.active
        # Check headers are present
        self.assertEqual(sheet.cell(row=1, column=1).value, "Date")
        self.assertGreater(sheet.cell(row=1, column=1).font.bold, False)

    def test_report_permission_denied(self) -> None:
        """Test that PG cannot generate reports for others."""
        self.client.force_authenticate(self.pg)
        url = reverse("reports_api:generate")
        response = self.client.post(
            url,
            {
                "template_slug": self.template.slug,
                "format": "pdf",
                "params": {"pg_id": self.supervisor.pk},  # Try to get supervisor's data
            },
            format="json",
        )
        # Should succeed but only show pg's own data
        self.assertEqual(response.status_code, 200)

    def test_invalid_format_raises_error(self) -> None:
        """Test that invalid format raises error."""
        service = ReportService(self.admin)
        with self.assertRaises(ValueError):
            service.generate(self.template, {}, "invalid")

    def test_scheduled_report_email_sent(self) -> None:
        """Test that scheduled report sends email."""
        from django.core import mail
        schedule = ScheduledReport.objects.create(
            template=self.template,
            created_by=self.admin,
            email_to="test@example.com",
            params={"format": "pdf"},
            cron="0 6 * * *",
            next_run_at=timezone.now(),
        )
        call_command("run_scheduled_reports")
        schedule.refresh_from_db()
        # Check email was sent
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("test@example.com", mail.outbox[0].to)
        self.assertEqual(len(mail.outbox[0].attachments), 1)

    def test_report_base64_encoding(self) -> None:
        """Test base64 encoding of report."""
        service = ReportService(self.admin)
        pdf_report = service.generate(self.template, {}, "pdf")
        base64_content = pdf_report.as_base64()
        # Decode and verify
        decoded = base64.b64decode(base64_content)
        self.assertEqual(decoded, pdf_report.content)

    def test_xlsx_column_widths(self) -> None:
        """Test that Excel columns have appropriate widths."""
        service = ReportService(self.admin)
        excel_report = service.generate(self.template, {}, "xlsx")
        workbook = load_workbook(io.BytesIO(excel_report.content))
        sheet = workbook.active
        # Check that columns have width set
        for column in sheet.column_dimensions.values():
            self.assertGreater(column.width, 10)

    def test_report_with_pg_filter(self) -> None:
        """Test report filtered by specific PG."""
        service = ReportService(self.admin)
        params = {"pg_id": str(self.pg.pk)}
        excel_report = service.generate(self.template, params, "xlsx")
        workbook = load_workbook(io.BytesIO(excel_report.content))
        sheet = workbook.active
        # Should have at least header + 1 row
        self.assertGreaterEqual(sheet.max_row, 2)

    def test_list_templates_api(self) -> None:
        """Test listing available report templates."""
        url = reverse("reports_api:templates")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertGreater(len(response.data), 0)

    def test_scheduled_report_failure_handling(self) -> None:
        """Test scheduled report handles failures."""
        # Create a schedule with invalid params (this will fail during generation)
        schedule = ScheduledReport.objects.create(
            template=self.template,
            created_by=self.admin,
            email_to="test@example.com",
            params={"format": "invalid_format"},  # Invalid format will cause failure
            cron="0 6 * * *",
            next_run_at=timezone.now(),
        )
        # This should not crash
        call_command("run_scheduled_reports")
        schedule.refresh_from_db()
        # Should record the failure
        self.assertIsNotNone(schedule.last_run_at)
        self.assertIn("error", schedule.last_result)
